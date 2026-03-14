"""
ConnectorOS Scout — Excel Generator (openpyxl)

Generates professional multi-sheet Excel deliverables:
  Sheet 1: Priority Leads (filtered, sorted)
  Sheet 2: Summary Stats
  Sheet 3: Needs Review (hot leads with shaky data)
  Sheet 4: Nurture List (good data, not urgently hiring)

Security:
  - File paths use pathlib (no shell injection)
  - Filenames sanitized to prevent directory traversal
  - Generated in EXPORTS_DIR only
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.config import EXPORTS_DIR
from core.logger import get_logger

logger = get_logger("export.excel")

# ── Colour Palette ───────────────────────────────────────────────
_RED_HOT_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
_WARM_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
_COOL_FILL = PatternFill(start_color="4488FF", end_color="4488FF", fill_type="solid")
_COLD_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

_VERIFIED_FILL = PatternFill(start_color="22CC44", end_color="22CC44", fill_type="solid")
_LIKELY_FILL = PatternFill(start_color="4488FF", end_color="4488FF", fill_type="solid")
_UNCERTAIN_FILL = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
_UNVERIFIED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

_HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Columns for the lead sheets
_LEAD_COLUMNS = [
    "Company", "Website", "Location", "Industry", "Employees",
    "Open Roles", "Top Roles", "Tech Stack", "Hiring Score",
    "Hiring Label", "Contact Name", "Contact Title", "Best Email",
    "LinkedIn", "Data Confidence", "Confidence Tier", "Notes",
]


def _sanitize_filename(name: str) -> str:
    """Remove unsafe characters from filenames."""
    return re.sub(r"[^\w\s\-.]", "", name).strip()


def _apply_header(ws, columns: list[str]) -> None:
    """Apply formatted header row."""
    for col_idx, title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"


def _apply_hiring_label_color(cell, label: str) -> None:
    """Colour-code hiring label cells."""
    fills = {
        "RED_HOT": _RED_HOT_FILL,
        "WARM": _WARM_FILL,
        "COOL": _COOL_FILL,
        "COLD": _COLD_FILL,
    }
    if label in fills:
        cell.fill = fills[label]
        if label in ("RED_HOT", "COLD"):
            cell.font = Font(color="FFFFFF", bold=True)


def _apply_confidence_color(cell, tier: str) -> None:
    """Colour-code confidence tier cells."""
    fills = {
        "VERIFIED": _VERIFIED_FILL,
        "LIKELY": _LIKELY_FILL,
        "UNCERTAIN": _UNCERTAIN_FILL,
        "UNVERIFIED": _UNVERIFIED_FILL,
    }
    if tier in fills:
        cell.fill = fills[tier]


def _write_lead_row(ws, row_num: int, lead: dict) -> None:
    """Write a single lead row (used across sheets)."""
    values = [
        lead.get("company_name", ""),
        lead.get("website_url", ""),
        lead.get("headquarters", ""),
        lead.get("industry", ""),
        lead.get("employee_count", ""),
        lead.get("role_count", 0),
        ", ".join(lead.get("top_roles", [])),
        ", ".join(lead.get("tech_stack", [])),
        lead.get("hiring_intensity", 0),
        lead.get("hiring_label", ""),
        lead.get("contact_name", ""),
        lead.get("contact_title", ""),
        lead.get("best_email", ""),
        lead.get("linkedin_url", ""),
        lead.get("data_confidence", 0),
        lead.get("confidence_tier", ""),
        lead.get("notes", ""),
    ]

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = _THIN_BORDER

        # Make URLs clickable
        if col_idx in (2, 14) and isinstance(value, str) and value.startswith("http"):
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")

        # Colour-code labels
        if col_idx == 10:  # Hiring Label
            _apply_hiring_label_color(cell, str(value))
        if col_idx == 16:  # Confidence Tier
            _apply_confidence_color(cell, str(value))


def _auto_width(ws) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)


def generate_excel(
    leads: list[dict],
    agency_name: str = "Default",
    min_hiring_for_priority: int = 60,
    min_confidence_for_priority: int = 60,
) -> str:
    """
    Generate a multi-sheet Excel file for agency delivery.

    Args:
        leads: List of lead dicts with all scoring/contact data.
        agency_name: Agency name (used in filename).
        min_hiring_for_priority: Min hiring score for priority sheet.
        min_confidence_for_priority: Min confidence for priority sheet.

    Returns:
        Absolute path to the generated Excel file.
    """
    wb = Workbook()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    safe_name = _sanitize_filename(agency_name)

    # ── Sheet 1: Priority Leads ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Priority Leads"
    _apply_header(ws1, _LEAD_COLUMNS)

    priority_leads = [
        l for l in leads
        if l.get("priority_tier") in ("PRIORITY", "REVIEW")
    ]
    priority_leads.sort(key=lambda x: x.get("hiring_intensity", 0), reverse=True)

    for i, lead in enumerate(priority_leads, 2):
        _write_lead_row(ws1, i, lead)
    _auto_width(ws1)

    # ── Sheet 2: Summary Stats ───────────────────────────────────
    ws2 = wb.create_sheet("Summary Stats")

    stats = [
        ("Total Leads", len(leads)),
        ("Priority Leads", len(priority_leads)),
        ("", ""),
        ("Hiring Label Breakdown", ""),
    ]

    # Count by hiring label
    for label in ("RED_HOT", "WARM", "COOL", "COLD"):
        count = sum(1 for l in leads if l.get("hiring_label") == label)
        pct = f"{count / max(len(leads), 1) * 100:.0f}%"
        stats.append((f"  {label}", f"{count} ({pct})"))

    stats.append(("", ""))
    stats.append(("Confidence Breakdown", ""))

    for tier in ("VERIFIED", "LIKELY", "UNCERTAIN", "UNVERIFIED"):
        count = sum(1 for l in leads if l.get("confidence_tier") == tier)
        pct = f"{count / max(len(leads), 1) * 100:.0f}%"
        stats.append((f"  {tier}", f"{count} ({pct})"))

    if leads:
        avg_hiring = sum(l.get("hiring_intensity", 0) for l in leads) / len(leads)
        avg_conf = sum(l.get("data_confidence", 0) for l in leads) / len(leads)
        stats.append(("", ""))
        stats.append(("Avg Hiring Score", f"{avg_hiring:.1f}"))
        stats.append(("Avg Data Confidence", f"{avg_conf:.1f}"))

    for row_idx, (label, value) in enumerate(stats, 1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=bool(label and not label.startswith(" ")))
        ws2.cell(row=row_idx, column=2, value=value)
    _auto_width(ws2)

    # ── Sheet 3: Needs Review ────────────────────────────────────
    ws3 = wb.create_sheet("Needs Review")
    _apply_header(ws3, _LEAD_COLUMNS)

    review_leads = [
        l for l in leads
        if l.get("hiring_intensity", 0) >= 60 and l.get("data_confidence", 0) < 60
    ]
    for i, lead in enumerate(review_leads, 2):
        _write_lead_row(ws3, i, lead)
    _auto_width(ws3)

    # ── Sheet 4: Nurture List ────────────────────────────────────
    ws4 = wb.create_sheet("Nurture List")
    _apply_header(ws4, _LEAD_COLUMNS)

    nurture_leads = [
        l for l in leads
        if l.get("hiring_intensity", 0) < 60 and l.get("data_confidence", 0) >= 60
    ]
    for i, lead in enumerate(nurture_leads, 2):
        _write_lead_row(ws4, i, lead)
    _auto_width(ws4)

    # ── Save ─────────────────────────────────────────────────────
    filename = f"ConnectorOS_{safe_name}_{today}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(str(filepath))

    logger.info(
        "excel_generated",
        file=str(filepath),
        total_leads=len(leads),
        priority=len(priority_leads),
        review=len(review_leads),
        nurture=len(nurture_leads),
    )

    return str(filepath)
